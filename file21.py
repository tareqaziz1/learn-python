# more Excel operations
import openpyxl as xl
from openpyxl.styles import PatternFill

path = "./others/transactions.xlsx"
workbook = xl.load_workbook(path)
sheet = workbook["Sheet1"]

for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 2)
    print(cell.value)


# Add a column as shipping cost

sheet.cell(row = 1, column = 4).value = "shipping_cost"

for row in range(2, sheet.max_row + 1):
    shipping = 5.5
    sheet.cell(row, 4).value = shipping

# Add a column as total cost. total cost will be - price + shipping cost

sheet.cell(row = 1, column = 5).value = "total_cost"

for row in range(2, sheet.max_row + 1):
    cell_price = sheet.cell(row, 3).value
    cell_shipping = sheet.cell(row, 4).value
    total_cost = cell_price + cell_shipping
    total_cost_cell = sheet.cell(row, 5)
    total_cost_cell.value = total_cost
    print(total_cost_cell.value)

# Make a value range based on price. Low (≤10), Medium (11–19), or High (≥20)

sheet.cell(row = 1, column = 6).value = "value_range"

for row in range(2, sheet.max_row + 1):
    price = sheet.cell(row, 3).value
    if price <= 10:
        value = "Low value"
    elif price <= 20:
        value = "Medium value"
    elif price >= 20:
        value = "High value"
    else:
        value = "none"

    sheet.cell(row, 6).value = value

# Make a column for profit. (Every product has 30% profit margin)

sheet.cell(row = 1, column = 7).value = "profit"

for row in range(2, sheet.max_row + 1):
    cell_profit = sheet.cell(row, 7)
    cell_price = sheet.cell(row, 3).value
    cell_profit.value = round(cell_price * 0.30, 2)            # 30% profit

from datetime import datetime

sheet.cell(row=1, column=8).value = "year"

for row in range(2, sheet.max_row + 1):
    sheet.cell(row, 8).value = datetime.now().year

for row in sheet.iter_rows(values_only=True):
    print(row)

# Style the rows with colors
#from openpyxl.styles import PatternFill

red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

for row in range(2, sheet.max_row + 1):
    price = sheet.cell(row, 3).value

    for col in range(5,7):
        if price > 20:
            sheet.cell(row, col).fill = green_fill
        else:
            sheet.cell(row, col).fill = red_fill





workbook.save('./others/transaction_updated_version2.xlsx')