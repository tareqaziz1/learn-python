# Charts with openpyxl

import openpyxl as xl
from openpyxl.chart import BarChart, LineChart, Reference

path = "./others/transactions_updated_v2.xlsx"
workbook = xl.load_workbook(path)
sheet = workbook["Sheet1"]

bar_chart = BarChart()  # Making a bar chart

data = Reference(sheet, min_col=3, min_row=1, max_row=sheet.max_row)  # Slicing the data
category = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row)    # Getting the label

bar_chart.add_data(data, titles_from_data=True)
bar_chart.set_categories(category)
bar_chart.title = "Price per product"
sheet.add_chart(bar_chart, "A25")  # Adding chart to the sheet in cell A25


# Making a line chart

chart = LineChart()

data = Reference(sheet, min_col=3, min_row=1, max_row=sheet.max_row)
cats = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

chart.title = "Price Trend"

sheet.add_chart(chart, "J38")


workbook.save("./others/transactions_updated_v3.xlsx")  # Save the workbook