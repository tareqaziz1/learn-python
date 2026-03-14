# Pypi and Pip
# installing openpyxl by pip install
# working with Excell sheets

import openpyxl as xl
from pathlib import Path

path = ("./others/transactions.xlsx")
workbook = xl.load_workbook(path)

sheet = workbook['Sheet1']

print(sheet.max_row) # 20

for row in range(1, sheet.max_row+1):
    cell = sheet.cell(row, 3)  # sheet.cell(row,column)
    print(cell.value)     # all the prices

# show all rows and columns
for row in sheet.iter_rows(values_only=True):
    print(row)

# To see as a table
'''
    for row in sheet.iter_rows(values_only=True):
    print("\t".join(str(cell) for cell in row))
'''

# Make a new row for 20% discounted price


sheet.cell(row=1, column=4).value = "discounted_price"      # add header for the new column

for row in range(2, sheet.max_row+1):
    cell = sheet.cell(row, 3)
    discounted_price = cell.value * 0.80  # 20% discount
    discounted_price_cell = sheet.cell(row, 4)
    discounted_price_cell.value = discounted_price
    print(discounted_price_cell.value)


workbook.save('./others/transaction_updated_version.xlsx')